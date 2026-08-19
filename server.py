import os
import re
import csv
import io
import json
import hashlib
import hmac
import secrets
import zipfile
from datetime import date as date_type, datetime, time as time_type, timedelta, timezone
from urllib.parse import urlparse, quote, unquote

from flask import Flask, request, jsonify, session, send_from_directory, Response
from flask_cors import CORS
from openpyxl import load_workbook
import pymysql
from pymysql.cursors import DictCursor

app = Flask(__name__, static_folder='.', static_url_path='')
IS_RENDER = bool(os.getenv('RENDER') or os.getenv('RENDER_SERVICE_ID'))
SECRET_KEY = os.getenv('SECRET_KEY')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD')
if IS_RENDER and not SECRET_KEY:
    raise RuntimeError('Render 部署缺少必要環境變數 SECRET_KEY')
if IS_RENDER and (not ADMIN_PASSWORD or ADMIN_PASSWORD == 'admin123'):
    raise RuntimeError('Render 部署必須設定非預設的 ADMIN_PASSWORD')
app.secret_key = SECRET_KEY or 'smart-ark-dev-secret'
CORS(app, supports_credentials=True)

DEFAULT_ADMIN = os.getenv('ADMIN_USERNAME', 'admin')
DEFAULT_PASSWORD = ADMIN_PASSWORD or 'admin123'
ROTATE_DEFAULT_ADMIN_PASSWORD = IS_RENDER and bool(ADMIN_PASSWORD)
DEFAULT_SHEET = os.getenv('ADMIN_DEFAULT_EVENT') or os.getenv('ADMIN_DEFAULT_EVENTS') or '活動報到名單'
CHECKED_STATUSES = ('checked_in', '已報到', '替代', 'done')
UTC_TZ = timezone.utc
TAIPEI_TZ = timezone(timedelta(hours=8), name='Asia/Taipei')


def utc_now_naive():
    """Return a timezone-free UTC value for MySQL DATETIME columns."""
    return datetime.now(UTC_TZ).replace(tzinfo=None)


def taipei_now():
    return datetime.now(TAIPEI_TZ)


def utc_db_datetime_to_taipei(value):
    """Interpret existing naive Render/MySQL values as UTC for display."""
    if not isinstance(value, datetime):
        return value
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC_TZ)
    return value.astimezone(TAIPEI_TZ)

# ============================================================
# DB helpers
# ============================================================

def db_params():
    url = os.getenv('DATABASE_URL') or os.getenv('MYSQL_URL')
    if url:
        u = urlparse(url)
        return dict(
            host=u.hostname,
            port=u.port or 3306,
            user=unquote(u.username or ''),
            password=unquote(u.password or ''),
            database=unquote((u.path or '').lstrip('/')) or os.getenv('MYSQLDATABASE'),
            charset='utf8mb4',
            cursorclass=DictCursor,
            autocommit=False,
            init_command="SET time_zone = '+00:00'",
        )
    return dict(
        host=os.getenv('MYSQLHOST') or os.getenv('DB_HOST') or 'localhost',
        port=int(os.getenv('MYSQLPORT') or os.getenv('DB_PORT') or 3306),
        user=os.getenv('MYSQLUSER') or os.getenv('DB_USER') or 'root',
        password=os.getenv('MYSQLPASSWORD') or os.getenv('DB_PASSWORD') or '',
        database=os.getenv('MYSQLDATABASE') or os.getenv('DB_NAME') or 'railway',
        charset='utf8mb4',
        cursorclass=DictCursor,
        autocommit=False,
        init_command="SET time_zone = '+00:00'",
    )


def get_db_connection():
    return pymysql.connect(**db_params())


def column_exists(cur, table, col):
    cur.execute(
        """
        SELECT COUNT(*) AS c
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
          AND COLUMN_NAME = %s
        """,
        (table, col),
    )
    return int((cur.fetchone() or {}).get('c') or 0) > 0


def table_columns(cur, table):
    cur.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table,),
    )
    return {r['COLUMN_NAME'] for r in cur.fetchall()}


def add_col(cur, table, col, spec):
    if not column_exists(cur, table, col):
        cur.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {spec}")


def try_sql(cur, sql, args=None):
    try:
        cur.execute(sql, args or ())
    except Exception:
        pass


def ensure_core_tables(conn):
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS admins (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(120) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                allowed_events LONGTEXT,
                current_event VARCHAR(255),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS event_configs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL DEFAULT 'admin',
                google_sheet_name VARCHAR(255) NOT NULL DEFAULT '活動報到名單',
                event_title VARCHAR(255),
                event_subtitle VARCHAR(255),
                date_start VARCHAR(80),
                date_end VARCHAR(80),
                brand_name VARCHAR(255),
                logo_url LONGTEXT,
                banner_image_url LONGTEXT,
                map_image_url LONGTEXT,
                products LONGTEXT,
                product_categories LONGTEXT,
                industry_mappings LONGTEXT,
                agenda LONGTEXT,
                event_config LONGTEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_event_config (admin_username, google_sheet_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS event_registrations (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL DEFAULT 'admin',
                google_sheet_name VARCHAR(255) NOT NULL DEFAULT '活動報到名單',
                admin_user VARCHAR(120) DEFAULT 'admin',
                event_key VARCHAR(255) DEFAULT '活動報到名單',
                name VARCHAR(255),
                phone VARCHAR(100),
                email VARCHAR(255),
                company VARCHAR(255),
                company_name VARCHAR(255),
                job_title VARCHAR(255),
                region VARCHAR(255),
                training_level VARCHAR(255),
                industry_category VARCHAR(255),
                meal_preference VARCHAR(80),
                seat VARCHAR(100),
                seating_chart VARCHAR(100),
                status VARCHAR(40) DEFAULT 'pending',
                is_original TINYINT(1) DEFAULT 1,
                proxy_name VARCHAR(255),
                proxy_phone VARCHAR(100),
                checked_in_at DATETIME NULL,
                checkin_time DATETIME NULL,
                portrait_consent TINYINT(1) NULL,
                portrait_consent_status VARCHAR(40),
                portrait_consent_time DATETIME NULL,
                special_notes LONGTEXT,
                note LONGTEXT,
                raw_data LONGTEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                KEY idx_event (admin_username, google_sheet_name),
                KEY idx_status (status)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS csv_import_previews (
                token_hash CHAR(64) PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL,
                google_sheet_name VARCHAR(255) NOT NULL,
                file_sha256 CHAR(64) NOT NULL,
                normalized_rows_sha256 CHAR(64) NOT NULL DEFAULT '',
                source_format VARCHAR(20) NOT NULL DEFAULT 'csv',
                worksheet_name VARCHAR(255),
                parser_version VARCHAR(50) NOT NULL DEFAULT 'registration-v1',
                valid_count INT NOT NULL,
                roster_count INT NOT NULL,
                roster_revision CHAR(64) NOT NULL,
                expires_at DATETIME NOT NULL,
                used_at DATETIME NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_csv_preview_target (admin_username, google_sheet_name),
                KEY idx_csv_preview_expiry (expires_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS agenda_items (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL DEFAULT 'admin',
                google_sheet_name VARCHAR(255) NOT NULL DEFAULT '活動報到名單',
                sort_order INT DEFAULT 0,
                time VARCHAR(60),
                event VARCHAR(255),
                title VARCHAR(255),
                description LONGTEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                KEY idx_agenda (admin_username, google_sheet_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        preview_cols = {
            'normalized_rows_sha256': "CHAR(64) NOT NULL DEFAULT ''",
            'source_format': "VARCHAR(20) NOT NULL DEFAULT 'csv'",
            'worksheet_name': 'VARCHAR(255)',
            'parser_version': "VARCHAR(50) NOT NULL DEFAULT 'registration-v1'",
        }
        for col, spec in preview_cols.items():
            add_col(cur, 'csv_import_previews', col, spec)

        registration_cols = {
            'industry_category': 'VARCHAR(255)',
            'meal_preference': 'VARCHAR(80)',
        }
        for col, spec in registration_cols.items():
            add_col(cur, 'event_registrations', col, spec)

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS industry_mappings (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL DEFAULT 'admin',
                google_sheet_name VARCHAR(255) NOT NULL DEFAULT '活動報到名單',
                company_name VARCHAR(255),
                keyword VARCHAR(255),
                category VARCHAR(255),
                industry VARCHAR(255),
                sort_order INT DEFAULT 0,
                KEY idx_industry (admin_username, google_sheet_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS exhibitors (
                id INT AUTO_INCREMENT PRIMARY KEY,
                admin_username VARCHAR(120) NOT NULL DEFAULT 'admin',
                google_sheet_name VARCHAR(255) NOT NULL DEFAULT '活動報到名單',
                name VARCHAR(255),
                company_name VARCHAR(255),
                industry VARCHAR(255),
                image_url LONGTEXT,
                logo LONGTEXT,
                website LONGTEXT,
                contact VARCHAR(255),
                description LONGTEXT,
                sort_order INT DEFAULT 0,
                KEY idx_exhibitor (admin_username, google_sheet_name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )

        cfg_cols = {
            'admin_username': "VARCHAR(120) NOT NULL DEFAULT 'admin'",
            'google_sheet_name': "VARCHAR(255) NOT NULL DEFAULT '活動報到名單'",
            'event_title': 'VARCHAR(255)',
            'event_subtitle': 'VARCHAR(255)',
            'event_date_start': 'VARCHAR(80)',
            'event_date_end': 'VARCHAR(80)',
            'card1_icon': 'VARCHAR(40)',
            'card1_title': 'VARCHAR(255)',
            'card1_subtitle': 'VARCHAR(255)',
            'card1_description': 'LONGTEXT',
            'card1_url': 'LONGTEXT',
            'card2_icon': 'VARCHAR(40)',
            'card2_title': 'VARCHAR(255)',
            'card2_subtitle': 'VARCHAR(255)',
            'card2_description': 'LONGTEXT',
            'card2_url': 'LONGTEXT',
            'card3_icon': 'VARCHAR(40)',
            'card3_title': 'VARCHAR(255)',
            'card3_subtitle': 'VARCHAR(255)',
            'card3_description': 'LONGTEXT',
            'card3_url': 'LONGTEXT',
            'card4_icon': 'VARCHAR(40)',
            'card4_title': 'VARCHAR(255)',
            'card4_subtitle': 'VARCHAR(255)',
            'card4_description': 'LONGTEXT',
            'card4_url': 'LONGTEXT',
            'products': 'LONGTEXT',
            'product_categories': 'LONGTEXT',
            'industry_mappings': 'LONGTEXT',
            'agenda': 'LONGTEXT',
            'event_config': 'LONGTEXT',
        }
        for col, spec in cfg_cols.items():
            add_col(cur, 'event_configs', col, spec)

        cur.execute("SELECT id FROM admins WHERE username=%s", (DEFAULT_ADMIN,))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO admins (username,password,allowed_events,current_event) VALUES (%s,%s,%s,%s)",
                (DEFAULT_ADMIN, DEFAULT_PASSWORD, DEFAULT_SHEET, DEFAULT_SHEET),
            )
        if ROTATE_DEFAULT_ADMIN_PASSWORD:
            cur.execute(
                "UPDATE admins SET password=%s WHERE password=%s",
                (DEFAULT_PASSWORD, 'admin123'),
            )
    conn.commit()


# ============================================================
# Generic helpers
# ============================================================

def get_payload():
    return request.get_json(silent=True) if request.is_json else None


def event_args():
    data = get_payload() or {}
    admin = (
        request.args.get('admin')
        or request.form.get('admin')
        or data.get('admin')
        or data.get('admin_username')
        or session.get('username')
        or DEFAULT_ADMIN
    )
    sheet = (
        request.args.get('sheet')
        or request.args.get('google_sheet_name')
        or request.args.get('event_key')
        or request.form.get('sheet')
        or data.get('sheet')
        or data.get('google_sheet_name')
        or session.get('current_admin_sheet')
        or DEFAULT_SHEET
    )
    return str(admin).strip() or DEFAULT_ADMIN, str(sheet).strip() or DEFAULT_SHEET


def q_event():
    return event_args()


def validate_admin_event_session(requested_admin, sheet):
    admin = clean_text(session.get('username'))
    if not session.get('admin_logged_in') or not admin:
        return None, (jsonify(success=False, message='請先登入管理員後台'), 401)
    if requested_admin != admin:
        return None, (jsonify(success=False, message='你沒有權限讀取這個管理員的資料'), 403)
    allowed = session.get('allowed_sheets') or []
    if isinstance(allowed, str):
        allowed = [value.strip() for value in allowed.split(',') if value.strip()]
    current = clean_text(session.get('current_admin_sheet'))
    allowed = {clean_text(value) for value in allowed if clean_text(value)}
    if current:
        allowed.add(current)
    if sheet not in allowed:
        return None, (jsonify(success=False, message='你沒有權限讀取這個場次'), 403)
    return admin, None


def json_loads(value, default=None):
    if default is None:
        default = []
    if value is None or value == '':
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


def json_dumps(value):
    return json.dumps(value if value is not None else [], ensure_ascii=False)


MAX_CSV_BYTES = 5 * 1024 * 1024
MAX_CSV_ROWS = 50000
MAX_XLSX_COLUMNS = 64
MAX_XLSX_CELL_CHARACTERS = 10000
MAX_XLSX_ZIP_ENTRIES = 1000
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 20 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
CSV_PREVIEW_TOKEN_TTL_SECONDS = 15 * 60
REGISTRATION_PARSER_VERSION = 'registration-v3'
REGISTRATION_FIELD_ALIASES = {
    'name': ['姓名', 'name', 'Name', '名字', '貴賓姓名'],
    'phone': ['手機', '電話', 'phone', 'Phone', '行動電話', '手機號碼'],
    'email': ['Email', 'email', 'E-mail', '信箱', '電子郵件'],
    'company': ['公司', '公司名稱', '公司/單位', '服務單位', '單位', 'company', 'Company'],
    'job_title': ['職稱', 'title', 'job_title', '職位', 'position', 'Position'],
    'region': ['地區', '區域', 'region', 'Region'],
    'training_level': ['職階', '層級', 'training_level', 'level'],
    'industry_category': ['產業類別', '產業', '行業類別', '行業', 'industry', 'industry_category'],
    'meal_preference': ['葷素', '餐別', '餐飲偏好', '飲食偏好', 'meal', 'meal_preference'],
    'seat': ['桌號', '桌號/座位', '座位', '桌次', 'seat', 'Seat', 'seating_chart'],
    'special_notes': ['備註', 'notes', 'note', 'Remarks'],
}
REGISTRATION_TEMPLATE_HEADERS = ['姓名', '手機', 'Email', '公司名稱', '職稱', '地區', '職階', '產業類別', '葷素', '桌號', '備註']
REGISTRATION_REQUIRED_FIELDS = {'name', 'phone', 'company'}


def normalize_csv_header(value):
    return str(value or '').lstrip('\ufeff').strip().casefold()


REGISTRATION_ALIAS_TO_FIELD = {
    normalize_csv_header(alias): field
    for field, aliases in REGISTRATION_FIELD_ALIASES.items()
    for alias in aliases
}


def normalize_csv_row(row):
    normalized = {}
    for key, value in (row or {}).items():
        header = normalize_csv_header(key)
        if header and header not in normalized:
            normalized[header] = value
    return normalized


def create_csv_preview_token():
    return secrets.token_urlsafe(32)


def csv_preview_token_hash(token):
    return hashlib.sha256(clean_text(token).encode('utf-8')).hexdigest()


def csv_file_hash(raw):
    return hashlib.sha256(raw).hexdigest()


def registration_rows_hash(rows):
    payload = json.dumps(
        rows or [],
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
    ).encode('utf-8')
    return hashlib.sha256(payload).hexdigest()


def csv_roster_revision(rows):
    canonical = json.dumps(
        rows or [],
        ensure_ascii=False,
        sort_keys=True,
        separators=(',', ':'),
        default=str,
    ).encode('utf-8')
    return hashlib.sha256(canonical).hexdigest()


def load_csv_roster_snapshot(cur, admin, sheet, for_update=False):
    sql = """
        SELECT * FROM event_registrations
        WHERE admin_username=%s AND google_sheet_name=%s
        ORDER BY id
    """
    if for_update:
        sql += " FOR UPDATE"
    cur.execute(sql, (admin, sheet))
    rows = cur.fetchall() or []
    return len(rows), csv_roster_revision(rows)


def lock_event_mutations(cur, admin, sheet):
    cur.execute(
        "SELECT id FROM event_configs WHERE admin_username=%s AND google_sheet_name=%s FOR UPDATE",
        (admin, sheet),
    )
    if not cur.fetchone():
        raise RuntimeError('找不到活動場次設定，請重新整理後再試')


def pick(row, keys):
    for k in keys:
        value = row.get(normalize_csv_header(k))
        if value is not None and str(value).strip():
            return str(value).strip()
    return ''


def clean_text(v):
    return str(v or '').strip()


def search_norm(v):
    text = str(v or '').lower().strip()
    text = re.sub(r'[\s\-_/\.、，,。．·・:：;；\(\)（）\[\]【】{}「」『』\+]+', '', text)
    return text


def search_tokens(v):
    raw = str(v or '').strip()
    tokens = [search_norm(x) for x in re.split(r'[\s、，,]+', raw) if search_norm(x)]
    whole = search_norm(raw)
    if whole and whole not in tokens:
        tokens.insert(0, whole)
    return tokens


def search_match_score(row, query, method='company'):
    q = search_norm(query)
    tokens = search_tokens(query)
    company = search_norm(row.get('company') or row.get('company_name'))
    name = search_norm(row.get('name'))
    phone = search_norm(row.get('phone'))
    email = search_norm(row.get('email'))
    job = search_norm(row.get('job_title'))
    hay_company = company
    hay_all = ''.join([company, name, phone, email, job])
    if not q and not tokens:
        return 0
    if method in ['company', 'company_name', 'unit']:
        if q and q in hay_company:
            return 100
        if tokens and all(t in hay_company for t in tokens):
            return 95
        if q and q in hay_all:
            return 80
        if tokens and all(t in hay_all for t in tokens):
            return 75
        if tokens and any(t in hay_company for t in tokens):
            return 50
    else:
        if q and q in hay_all:
            return 90
        if tokens and all(t in hay_all for t in tokens):
            return 80
    return 0


def table_sort_key(label):
    s = clean_text(label)
    nums = re.findall(r'\d+', s)
    return (0, int(nums[0])) if nums else (1, s)


def normalize_table_label(value):
    v = clean_text(value)
    if not v:
        return ''
    v = v.replace('桌', '').replace('第', '').strip()
    return v or clean_text(value)


def normalize_industry_category(value):
    return clean_text(value)


def normalize_meal_preference(value):
    raw = clean_text(value)
    if not raw:
        return '不用餐'
    key = re.sub(r'\s+', '', raw).casefold()
    aliases = {
        '葷': '葷食',
        '葷食': '葷食',
        '一般餐': '葷食',
        '非素食': '葷食',
        '素': '素食',
        '素食': '素食',
        '蛋奶素': '蛋奶素',
        '奶蛋素': '蛋奶素',
        '全素': '全素',
        '純素': '全素',
        'vegan': '全素',
        '不用餐': '不用餐',
        '不需餐': '不用餐',
        '不需要餐食': '不用餐',
        '不需餐食': '不用餐',
        '無需餐食': '不用餐',
    }
    return aliases.get(key, raw)


def status_checked(status):
    return str(status or '').lower() in ['checked_in', '已報到', '替代', 'done']


def is_vegetarian_meal(value):
    return normalize_meal_preference(value) in {'素食', '蛋奶素', '全素'}


def normalize_registration(row):
    normalized = normalize_csv_row(row)
    return {
        'name': pick(normalized, REGISTRATION_FIELD_ALIASES['name']),
        'phone': pick(normalized, REGISTRATION_FIELD_ALIASES['phone']),
        'email': pick(normalized, REGISTRATION_FIELD_ALIASES['email']),
        'company': pick(normalized, REGISTRATION_FIELD_ALIASES['company']),
        'job_title': pick(normalized, REGISTRATION_FIELD_ALIASES['job_title']),
        'region': pick(normalized, REGISTRATION_FIELD_ALIASES['region']),
        'training_level': pick(normalized, REGISTRATION_FIELD_ALIASES['training_level']),
        'industry_category': normalize_industry_category(pick(normalized, REGISTRATION_FIELD_ALIASES['industry_category'])),
        'meal_preference': normalize_meal_preference(pick(normalized, REGISTRATION_FIELD_ALIASES['meal_preference'])),
        'seat': normalize_table_label(pick(normalized, REGISTRATION_FIELD_ALIASES['seat'])),
        'special_notes': pick(normalized, REGISTRATION_FIELD_ALIASES['special_notes']),
        'raw_data': json_dumps(row),
    }


def validate_registration_headers(fieldnames, format_label):
    cleaned = [str(name or '').lstrip('\ufeff').strip() for name in (fieldnames or [])]
    if not cleaned or not any(cleaned):
        raise ValueError(f'{format_label} 沒有標題列，請確認第一列是欄位名稱')

    recognized_columns = [
        REGISTRATION_ALIAS_TO_FIELD[normalize_csv_header(name)]
        for name in cleaned
        if normalize_csv_header(name) in REGISTRATION_ALIAS_TO_FIELD
    ]
    recognized_fields = set(recognized_columns)
    if not recognized_fields:
        raise ValueError(f'{format_label} 欄名無法辨識，請使用系統提供的範本')
    duplicate_fields = sorted({field for field in recognized_columns if recognized_columns.count(field) > 1})
    if duplicate_fields:
        raise ValueError(f'{format_label} 有重複用途的欄位，請使用系統提供的範本')
    if not recognized_fields.intersection(REGISTRATION_REQUIRED_FIELDS):
        raise ValueError(f'{format_label} 至少需要「姓名」、「手機」或「公司名稱」其中一個欄位')
    return cleaned, recognized_fields


def registration_parse_result(format_label, source_format, headers, recognized_fields, valid_rows, total_count, **extra):
    skipped_count = total_count - len(valid_rows)
    if not valid_rows:
        raise ValueError(f'{format_label} 沒有可匯入的有效資料，原有名單未變更')
    return {
        'source_format': source_format,
        'headers': headers,
        'recognized_fields': sorted(recognized_fields),
        'rows': valid_rows,
        'total_count': total_count,
        'valid_count': len(valid_rows),
        'skipped_count': skipped_count,
        **extra,
    }


def decode_registration_csv(raw):
    if not raw:
        raise ValueError('CSV 檔案是空的')
    if b'\x00' in raw:
        raise ValueError('CSV 看起來是 UTF-16 格式，請在 Excel 另存為「CSV UTF-8」')

    last_error = None
    for encoding in ['utf-8-sig', 'utf-8', 'cp950', 'big5', 'big5hkscs']:
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError as err:
            last_error = err
    raise ValueError(f'CSV 編碼解析失敗，請另存為「CSV UTF-8」：{last_error}')


def parse_registration_csv(raw):
    text, encoding = decode_registration_csv(raw)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel

    try:
        extra_columns_key = '__extra_columns__'
        reader = csv.DictReader(
            io.StringIO(text, newline=''),
            dialect=dialect,
            restkey=extra_columns_key,
            restval=None,
            strict=True,
        )
        fieldnames, recognized_fields = validate_registration_headers(reader.fieldnames, 'CSV')

        valid_rows = []
        total_count = 0
        for row_number, row in enumerate(reader, start=1):
            if row_number > MAX_CSV_ROWS:
                raise ValueError(f'CSV 超過 {MAX_CSV_ROWS:,} 筆上限，請拆成較小的檔案')
            if row.get(extra_columns_key) is not None:
                raise ValueError(f'CSV 第 {reader.line_num} 列欄位數超過表頭')
            if any(value is None for key, value in row.items() if key != extra_columns_key):
                raise ValueError(f'CSV 第 {reader.line_num} 列欄位數少於表頭')
            row.pop(extra_columns_key, None)
            total_count += 1
            normalized_row = normalize_registration(row)
            if normalized_row['name'] or normalized_row['phone'] or normalized_row['company']:
                valid_rows.append(normalized_row)
    except csv.Error as err:
        raise ValueError(f'CSV 格式解析失敗：{err}') from err

    return registration_parse_result(
        'CSV', 'csv', fieldnames, recognized_fields, valid_rows, total_count,
        encoding=encoding,
        worksheet_name=None,
    )


def validate_xlsx_archive(raw):
    if not raw:
        raise ValueError('Excel 檔案是空的')
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            members = archive.infolist()
            if not members or len(members) > MAX_XLSX_ZIP_ENTRIES:
                raise ValueError('Excel 檔案結構過大或不完整')

            names = [member.filename for member in members]
            if len(names) != len(set(names)):
                raise ValueError('Excel 檔案含有重複的內部項目')
            if '[Content_Types].xml' not in names or 'xl/workbook.xml' not in names:
                raise ValueError('檔案不是有效的 XLSX 活頁簿')
            lower_names = [name.lower() for name in names]
            if any(
                name.startswith(('xl/vbaproject', 'xl/activex/', 'xl/embeddings/'))
                for name in lower_names
            ):
                raise ValueError('不支援含巨集、ActiveX 或嵌入物件的 Excel 檔案')

            total_size = 0
            for member in members:
                name = member.filename.replace('\\', '/')
                path_parts = [part for part in name.split('/') if part not in ['', '.']]
                if (
                    '\x00' in name
                    or name.startswith('/')
                    or re.match(r'^[A-Za-z]:', name)
                    or '..' in path_parts
                    or member.flag_bits & 0x1
                    or member.compress_type not in [zipfile.ZIP_STORED, zipfile.ZIP_DEFLATED]
                ):
                    raise ValueError('Excel 檔案含有不安全或不支援的內部項目')
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ValueError('Excel 檔案解壓後的單一項目過大')
                if member.compress_size == 0:
                    ratio = float('inf') if member.file_size else 1
                else:
                    ratio = member.file_size / member.compress_size
                if ratio > MAX_XLSX_COMPRESSION_RATIO:
                    raise ValueError('Excel 檔案壓縮比例異常，無法安全處理')
                total_size += member.file_size
                if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError('Excel 檔案解壓後超過安全上限')
    except (zipfile.BadZipFile, zipfile.LargeZipFile) as err:
        raise ValueError('檔案不是有效的 XLSX 活頁簿') from err


def excel_cell_to_text(cell, field=None):
    if cell.data_type == 'f':
        raise ValueError(f'Excel 第 {cell.row} 列「{cell.column_letter}」含公式，請先貼上為純值')
    if cell.data_type == 'e':
        raise ValueError(f'Excel 第 {cell.row} 列「{cell.column_letter}」含錯誤值，請先修正')

    value = cell.value
    if value is None:
        return ''
    if isinstance(value, bool):
        raise ValueError(f'Excel 第 {cell.row} 列「{cell.column_letter}」是布林值，請改成文字')
    if isinstance(value, (datetime, date_type, time_type)):
        if field == 'phone':
            raise ValueError(f'Excel 第 {cell.row} 列手機不是文字格式，請改為文字以保留原始內容')
        text = value.isoformat(sep=' ') if isinstance(value, datetime) else value.isoformat()
    elif isinstance(value, int):
        if field == 'phone':
            raise ValueError(f'Excel 第 {cell.row} 列手機是數字格式，請改為文字以保留開頭的 0')
        text = str(value)
    elif isinstance(value, float):
        if value != value or value in [float('inf'), float('-inf')]:
            raise ValueError(f'Excel 第 {cell.row} 列「{cell.column_letter}」不是有效數字')
        if field == 'phone':
            raise ValueError(f'Excel 第 {cell.row} 列手機是數字格式，請改為文字以保留開頭的 0')
        text = str(int(value)) if value.is_integer() else format(value, '.15g')
    else:
        text = str(value).strip()
    if len(text) > MAX_XLSX_CELL_CHARACTERS:
        raise ValueError(f'Excel 第 {cell.row} 列「{cell.column_letter}」文字超過長度上限')
    return text


def find_registration_worksheet(workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != 'visible':
            continue
        merged_cells = getattr(worksheet, 'merged_cells', None)
        if merged_cells is not None and len(merged_cells.ranges):
            raise ValueError(f'Excel 工作表「{worksheet.title}」含合併儲存格，請取消合併後再匯入')
        reported_columns = worksheet.max_column or 0
        scan_columns = min(max(reported_columns, 1), MAX_XLSX_COLUMNS + 1)
        header_cells = next(
            worksheet.iter_rows(min_row=1, max_row=1, max_col=scan_columns),
            (),
        )
        headers = [excel_cell_to_text(cell) for cell in header_cells]
        recognized = {
            REGISTRATION_ALIAS_TO_FIELD[normalize_csv_header(header)]
            for header in headers
            if normalize_csv_header(header) in REGISTRATION_ALIAS_TO_FIELD
        }
        if recognized.intersection(REGISTRATION_REQUIRED_FIELDS):
            if reported_columns > MAX_XLSX_COLUMNS:
                raise ValueError(f'Excel 工作表「{worksheet.title}」欄位超過 {MAX_XLSX_COLUMNS} 欄上限')
            candidates.append((worksheet, headers))

    if not candidates:
        raise ValueError('Excel 找不到可辨識的報到名單工作表；請把欄位名稱放在第一列')
    if len(candidates) > 1:
        names = '、'.join(f'「{worksheet.title}」' for worksheet, _ in candidates[:5])
        raise ValueError(f'Excel 有多張可匯入的工作表（{names}），請只保留一張報到名單')
    return candidates[0]


def parse_registration_xlsx(raw):
    validate_xlsx_archive(raw)
    workbook = None
    try:
        workbook = load_workbook(
            io.BytesIO(raw),
            read_only=True,
            data_only=False,
            keep_links=False,
            keep_vba=False,
            rich_text=False,
        )
        worksheet, header_values = find_registration_worksheet(workbook)
        fieldnames, recognized_fields = validate_registration_headers(header_values, 'Excel')
        field_by_index = [REGISTRATION_ALIAS_TO_FIELD.get(normalize_csv_header(name)) for name in fieldnames]
        valid_rows = []
        total_count = 0
        for scanned_rows, cells in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(fieldnames)),
            start=1,
        ):
            if scanned_rows > MAX_CSV_ROWS:
                raise ValueError(f'Excel 超過 {MAX_CSV_ROWS:,} 筆上限，請拆成較小的檔案')
            row_values = [
                excel_cell_to_text(cell, field_by_index[index])
                for index, cell in enumerate(cells)
            ]
            if not any(row_values):
                continue
            total_count += 1
            raw_row = dict(zip(fieldnames, row_values))
            normalized_row = normalize_registration(raw_row)
            if normalized_row['name'] or normalized_row['phone'] or normalized_row['company']:
                valid_rows.append(normalized_row)

        return registration_parse_result(
            'Excel', 'xlsx', fieldnames, recognized_fields, valid_rows, total_count,
            encoding=None,
            worksheet_name=worksheet.title,
        )
    except ValueError:
        raise
    except Exception as err:
        raise ValueError('檔案不是有效的 XLSX 活頁簿') from err
    finally:
        if workbook is not None:
            workbook.close()


def parse_registration_upload(raw, filename):
    extension = os.path.splitext(clean_text(filename).lower())[1]
    if extension == '.csv':
        if raw.startswith(b'PK\x03\x04'):
            raise ValueError('檔案內容是 Excel，請使用正確的 .xlsx 副檔名')
        return parse_registration_csv(raw)
    if extension == '.xlsx':
        return parse_registration_xlsx(raw)
    if extension in ['.xls', '.xlsm', '.xlsb']:
        raise ValueError('只支援 .csv 或 .xlsx；請將舊版或含巨集的 Excel 另存為 .xlsx')
    raise ValueError('只支援 .csv 或 .xlsx 報到名單')


def portrait_status_from_row(row):
    raw = clean_text((row or {}).get('portrait_consent_status'))
    if raw:
        return raw
    v = (row or {}).get('portrait_consent')
    if v in [1, True, '1', 'true', 'True', '同意', 'yes', 'Yes']:
        return '同意'
    if v in [0, False, '0', 'false', 'False', '不同意', 'no', 'No']:
        return '不同意'
    return '未填'


def public_user(row):
    r = dict(row or {})
    company = r.get('company') or r.get('company_name') or ''
    seat = r.get('seat') or r.get('seating_chart') or ''
    checked_time = utc_db_datetime_to_taipei(r.get('checked_in_at') or r.get('checkin_time'))
    portrait_time = utc_db_datetime_to_taipei(r.get('portrait_consent_time'))
    portrait_status = portrait_status_from_row(r)
    original_name = clean_text(r.get('name'))
    original_phone = clean_text(r.get('phone'))
    original_email = clean_text(r.get('email'))
    proxy_name = clean_text(r.get('proxy_name'))
    proxy_phone = clean_text(r.get('proxy_phone'))
    proxy_email = clean_text(r.get('proxy_email'))
    is_proxy = str(r.get('status') or '').lower() == '替代' or r.get('is_original') in (0, False, '0')
    attendance_name = proxy_name if is_proxy and proxy_name else original_name
    attendance_phone = proxy_phone if is_proxy and proxy_name else original_phone
    attendance_email = proxy_email if is_proxy and proxy_name else original_email
    meal_preference = normalize_meal_preference(r.get('meal_preference'))
    result = {
        **r,
        'original_name': original_name,
        'original_phone': original_phone,
        'original_email': original_email,
        'proxy_name': proxy_name,
        'proxy_phone': proxy_phone,
        'proxy_email': proxy_email,
        'attendance_name': attendance_name,
        'attendance_phone': attendance_phone,
        'attendance_email': attendance_email,
        'display_name': attendance_name,
        'is_proxy': bool(is_proxy and proxy_name),
        'company': company,
        'company_name': company,
        'seat': seat,
        'table': seat,
        'seating_chart': seat,
        'job_title': r.get('job_title') or '',
        'industry_category': r.get('industry_category') or '',
        'industry': r.get('industry_category') or '',
        'meal_preference': meal_preference,
        'meal': meal_preference,
        'portrait_consent_status': portrait_status,
        'portrait_consent': True if portrait_status == '同意' else (False if portrait_status == '不同意' else None),
        'checked_in_at': checked_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(checked_time, 'strftime') else (checked_time or ''),
        'checkin_time': checked_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(checked_time, 'strftime') else (checked_time or ''),
        'checkedInAt': checked_time.strftime('%H:%M:%S') if hasattr(checked_time, 'strftime') else (checked_time or ''),
        'checked_in_at_iso': checked_time.isoformat(timespec='seconds') if isinstance(checked_time, datetime) else '',
        'portrait_consent_time': portrait_time.strftime('%Y-%m-%d %H:%M:%S') if hasattr(portrait_time, 'strftime') else (portrait_time or ''),
        'portrait_consent_time_iso': portrait_time.isoformat(timespec='seconds') if isinstance(portrait_time, datetime) else '',
        'special_notes': r.get('special_notes') or r.get('note') or '',
        'note': r.get('special_notes') or r.get('note') or '',
    }
    return result


def ensure_config(conn, admin, sheet):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM event_configs WHERE admin_username=%s AND google_sheet_name=%s LIMIT 1",
            (admin, sheet),
        )
        if not cur.fetchone():
            cur.execute(
                """
                INSERT INTO event_configs (admin_username, google_sheet_name, event_title, event_subtitle, brand_name)
                VALUES (%s,%s,%s,%s,%s)
                """,
                (admin, sheet, sheet, '世代共榮的數位聚合', '智匯方舟'),
            )
    conn.commit()


def get_config_row(conn, admin, sheet):
    ensure_config(conn, admin, sheet)
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM event_configs WHERE admin_username=%s AND google_sheet_name=%s LIMIT 1", (admin, sheet))
        return cur.fetchone() or {}


def load_agenda(conn, admin, sheet):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT time, COALESCE(NULLIF(title,''), event) AS title, event, description, sort_order FROM agenda_items WHERE admin_username=%s AND google_sheet_name=%s ORDER BY sort_order ASC, id ASC",
            (admin, sheet),
        )
        rows = cur.fetchall()
    if rows:
        return [dict(r) for r in rows]
    cfg = get_config_row(conn, admin, sheet)
    return json_loads(cfg.get('agenda'), [])


def save_agenda(conn, admin, sheet, rows):
    rows = rows or []
    ensure_config(conn, admin, sheet)
    with conn.cursor() as cur:
        cur.execute("DELETE FROM agenda_items WHERE admin_username=%s AND google_sheet_name=%s", (admin, sheet))
        for i, item in enumerate(rows):
            time = clean_text(item.get('time'))
            title = clean_text(item.get('title') or item.get('event'))
            desc = clean_text(item.get('description'))
            if not (time or title or desc):
                continue
            cur.execute(
                """
                INSERT INTO agenda_items (admin_username, google_sheet_name, sort_order, time, event, title, description)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                """,
                (admin, sheet, int(item.get('sort_order', i) or i), time, title, title, desc),
            )
        cur.execute("UPDATE event_configs SET agenda=%s WHERE admin_username=%s AND google_sheet_name=%s", (json_dumps(rows), admin, sheet))
    conn.commit()


def load_industry_mappings(conn, admin, sheet):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT company_name, keyword, category, industry, sort_order FROM industry_mappings WHERE admin_username=%s AND google_sheet_name=%s ORDER BY sort_order ASC, id ASC",
            (admin, sheet),
        )
        rows = cur.fetchall()
    if rows:
        out = []
        for r in rows:
            keyword = r.get('keyword') or r.get('company_name') or ''
            category = r.get('category') or r.get('industry') or '其他'
            out.append({**r, 'keyword': keyword, 'category': category, 'company': r.get('company_name') or keyword})
        return out
    cfg = get_config_row(conn, admin, sheet)
    return json_loads(cfg.get('industry_mappings'), [])


def save_industry_mappings(conn, admin, sheet, rows):
    rows = rows or []
    ensure_config(conn, admin, sheet)
    with conn.cursor() as cur:
        cur.execute("DELETE FROM industry_mappings WHERE admin_username=%s AND google_sheet_name=%s", (admin, sheet))
        for i, item in enumerate(rows):
            company = clean_text(item.get('company_name') or item.get('company') or item.get('keyword'))
            keyword = clean_text(item.get('keyword') or company)
            category = clean_text(item.get('category') or item.get('industry')) or '其他'
            if not (company or keyword or category):
                continue
            cur.execute(
                """
                INSERT INTO industry_mappings (admin_username, google_sheet_name, company_name, keyword, category, industry, sort_order)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                """,
                (admin, sheet, company, keyword, category, category, int(item.get('sort_order', i) or i)),
            )
        cur.execute("UPDATE event_configs SET industry_mappings=%s WHERE admin_username=%s AND google_sheet_name=%s", (json_dumps(rows), admin, sheet))
    conn.commit()


def load_exhibitors(conn, admin, sheet):
    with conn.cursor() as cur:
        cur.execute(
            "SELECT name, company_name, industry, image_url, logo, website, contact, description, sort_order FROM exhibitors WHERE admin_username=%s AND google_sheet_name=%s ORDER BY sort_order ASC, id ASC",
            (admin, sheet),
        )
        rows = cur.fetchall()
    out = []
    for r in rows:
        name = r.get('name') or r.get('company_name') or ''
        image = r.get('image_url') or r.get('logo') or ''
        out.append({**r, 'name': name, 'company_name': name, 'image_url': image, 'logo': image})
    return out


def save_exhibitors(conn, admin, sheet, rows):
    rows = rows or []
    with conn.cursor() as cur:
        cur.execute("DELETE FROM exhibitors WHERE admin_username=%s AND google_sheet_name=%s", (admin, sheet))
        for i, item in enumerate(rows):
            name = clean_text(item.get('name') or item.get('company_name'))
            if not name:
                continue
            image = clean_text(item.get('image_url') or item.get('logo'))
            cur.execute(
                """
                INSERT INTO exhibitors (admin_username, google_sheet_name, name, company_name, industry, image_url, logo, website, contact, description, sort_order)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    admin, sheet, name, name,
                    clean_text(item.get('industry')),
                    image, image,
                    clean_text(item.get('website')),
                    clean_text(item.get('contact')),
                    clean_text(item.get('description')),
                    int(item.get('sort_order', i) or i),
                ),
            )
    conn.commit()


def load_products_from_config(row):
    products = json_loads((row or {}).get('products'), [])
    if isinstance(products, dict):
        products = products.get('items') or products.get('products') or []
    return products if isinstance(products, list) else []


def save_config(conn, admin, sheet, payload):
    payload = dict(payload or {})
    ensure_config(conn, admin, sheet)
    cols = None
    with conn.cursor() as cur:
        cols = table_columns(cur, 'event_configs')
        direct = {}
        extra = {}
        for k, v in payload.items():
            if k in ['products', 'industry_mappings', 'agenda', 'product_categories']:
                direct[k] = json_dumps(v)
            elif k in cols and k not in ['id', 'created_at', 'updated_at', 'success_card_config', 'success_info_cards_config', 'dashboard_agenda_config']:
                direct[k] = v
            else:
                extra[k] = v
        if extra:
            row = get_config_row(conn, admin, sheet)
            old = json_loads(row.get('event_config'), {})
            if not isinstance(old, dict):
                old = {}
            old.update(extra)
            direct['event_config'] = json_dumps(old)
        if direct:
            set_sql = ', '.join([f"`{k}`=%s" for k in direct])
            cur.execute(f"UPDATE event_configs SET {set_sql} WHERE admin_username=%s AND google_sheet_name=%s", list(direct.values()) + [admin, sheet])
    conn.commit()


def serialize_config(row, conn=None, admin=None, sheet=None):
    cfg = dict(row or {})
    extra = json_loads(cfg.get('event_config'), {})
    if isinstance(extra, dict):
        cfg.update({k: v for k, v in extra.items() if k not in cfg or cfg.get(k) in [None, '']})
    if cfg.get('show_meal_options') is None:
        cfg['show_meal_options'] = True
    cfg['google_sheet_name'] = cfg.get('google_sheet_name') or sheet or DEFAULT_SHEET
    cfg['admin_username'] = cfg.get('admin_username') or admin or DEFAULT_ADMIN
    cfg['event_title'] = cfg.get('event_title') or cfg.get('exp_event_title') or cfg['google_sheet_name']
    cfg['event_subtitle'] = cfg.get('event_subtitle') or cfg.get('exp_event_subtitle') or '世代共榮的數位聚合'
    cfg['date_start'] = cfg.get('date_start') or cfg.get('event_date_start') or cfg.get('exp_event_date_start') or ''
    cfg['date_end'] = cfg.get('date_end') or cfg.get('event_date_end') or cfg.get('exp_event_date_end') or ''
    cfg['brand_name'] = cfg.get('brand_name') or cfg.get('exp_brand_name') or '智匯方舟'
    cfg['products'] = load_products_from_config(cfg)
    cfg['product_categories'] = json_loads(cfg.get('product_categories'), ['課程', '諮詢', '聯盟', '紀念品'])
    cfg['industry_mappings'] = json_loads(cfg.get('industry_mappings'), [])
    cfg['agenda'] = json_loads(cfg.get('agenda'), [])
    if conn and admin and sheet:
        table_agenda = load_agenda(conn, admin, sheet)
        table_mapping = load_industry_mappings(conn, admin, sheet)
        table_exhibitors = load_exhibitors(conn, admin, sheet)
        if table_agenda:
            cfg['agenda'] = table_agenda
        if table_mapping:
            cfg['industry_mappings'] = table_mapping
        cfg['exhibitors'] = table_exhibitors
    for k in list(cfg.keys()):
        if isinstance(cfg[k], datetime):
            cfg[k] = cfg[k].strftime('%Y-%m-%d %H:%M:%S')
    return cfg


def get_logs(conn, admin, sheet, limit=None, checked_only=False):
    sql = """
        SELECT * FROM event_registrations
        WHERE admin_username=%s AND google_sheet_name=%s
    """
    args = [admin, sheet]
    if checked_only:
        sql += " AND status IN ('checked_in','已報到','替代','done')"
    sql += " ORDER BY COALESCE(checked_in_at, checkin_time, created_at) DESC, id DESC"
    if limit:
        sql += " LIMIT %s"
        args.append(int(limit))
    with conn.cursor() as cur:
        cur.execute(sql, args)
        return [public_user(r) for r in cur.fetchall()]


# ============================================================
# Pages
# ============================================================
@app.route('/')
def index():
    return send_from_directory('.', '活動報到系統.html')


@app.route('/admin')
def admin_page():
    return send_from_directory('.', 'admin.html')


@app.route('/dashboard')
def dashboard_page():
    return send_from_directory('.', 'dashboard.html')


@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)


# ============================================================
# Auth / health
# ============================================================
@app.route('/api/health')
def health():
    now = taipei_now()
    return jsonify(
        success=True,
        status='ok',
        time=now.strftime('%Y-%m-%d %H:%M:%S'),
        time_iso=now.isoformat(timespec='seconds'),
        timezone='Asia/Taipei',
    )

@app.route('/api/login', methods=['POST'])
def api_login():
    conn = None
    try:
        data = get_payload() or {}
        username = clean_text(data.get('username'))
        password = clean_text(data.get('password'))
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM admins WHERE TRIM(username)=%s AND TRIM(password)=%s LIMIT 1", (username, password))
            admin = cur.fetchone()
        if not admin:
            return jsonify(success=False, message='帳密錯誤'), 401
        allowed = [x.strip() for x in (admin.get('allowed_events') or DEFAULT_SHEET).split(',') if x.strip()] or [DEFAULT_SHEET]
        current = admin.get('current_event') if admin.get('current_event') in allowed else allowed[0]
        session['admin_logged_in'] = True
        session['username'] = admin['username']
        session['allowed_sheets'] = allowed
        session['current_admin_sheet'] = current
        return jsonify(success=True, username=admin['username'], allowed_sheets=allowed, sheets=allowed, current_sheet=current)
    except Exception as e:
        return jsonify(success=False, message=f'登入 API 失敗：{e}'), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/logout')
def api_logout():
    session.clear()
    return jsonify(success=True)

# ============================================================
# Sheets / event selection
# ============================================================
@app.route('/api/current_sheet')
def current_sheet():
    admin = clean_text(request.args.get('admin')) or session.get('username') or DEFAULT_ADMIN
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT current_event, allowed_events FROM admins WHERE username=%s LIMIT 1", (admin,))
            row = cur.fetchone() or {}
        allowed = [x.strip() for x in (row.get('allowed_events') or DEFAULT_SHEET).split(',') if x.strip()] or [DEFAULT_SHEET]
        current = row.get('current_event') or session.get('current_admin_sheet') or allowed[0]
        if current not in allowed:
            current = allowed[0]
        return jsonify(success=True, current_sheet=current, sheet=current, sheets=allowed)
    except Exception as e:
        fallback = session.get('current_admin_sheet') or DEFAULT_SHEET
        return jsonify(success=True, current_sheet=fallback, sheet=fallback, warning=str(e))
    finally:
        if conn:
            conn.close()


@app.route('/api/sheets/list')
def sheets_list():
    admin = clean_text(session.get('username'))
    if not session.get('admin_logged_in') or not admin:
        return jsonify(success=False, message='請先登入管理員後台'), 401
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT allowed_events, current_event FROM admins WHERE username=%s LIMIT 1", (admin,))
            row = cur.fetchone() or {}
            allowed = [x.strip() for x in (row.get('allowed_events') or DEFAULT_SHEET).split(',') if x.strip()] or [DEFAULT_SHEET]
            cur.execute("SELECT DISTINCT google_sheet_name FROM event_registrations WHERE admin_username=%s", (admin,))
            from_data = [r['google_sheet_name'] for r in cur.fetchall() if r.get('google_sheet_name')]
            cur.execute("SELECT DISTINCT google_sheet_name FROM event_configs WHERE admin_username=%s", (admin,))
            from_cfg = [r['google_sheet_name'] for r in cur.fetchall() if r.get('google_sheet_name')]
        sheets = []
        for s in allowed + from_cfg + from_data:
            if s and s not in sheets:
                sheets.append(s)
        current = row.get('current_event') or session.get('current_admin_sheet') or (sheets[0] if sheets else DEFAULT_SHEET)
        return jsonify(
            success=True,
            username=admin,
            sheets=sheets or [DEFAULT_SHEET],
            allowed_sheets=sheets or [DEFAULT_SHEET],
            current_sheet=current,
        )
    except Exception as e:
        return jsonify(success=False, message=str(e), sheets=[]), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/session/sheet', methods=['POST'])
def session_sheet():
    conn = None
    try:
        data = get_payload() or {}
        admin = clean_text(session.get('username'))
        if not session.get('admin_logged_in') or not admin:
            return jsonify(success=False, message='請先登入管理員後台'), 401
        sheet = clean_text(data.get('sheet') or data.get('google_sheet_name') or data.get('event_key')) or DEFAULT_SHEET
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT allowed_events FROM admins WHERE username=%s LIMIT 1", (admin,))
            row = cur.fetchone()
            if not row:
                return jsonify(success=False, message='管理員帳號不存在，請重新登入'), 403
            allowed = [x.strip() for x in (row.get('allowed_events') or '').split(',') if x.strip()]
            if sheet not in allowed:
                allowed.append(sheet)
            cur.execute("UPDATE admins SET current_event=%s, allowed_events=%s WHERE username=%s", (sheet, ','.join(allowed), admin))
        ensure_config(conn, admin, sheet)
        conn.commit()
        session['allowed_sheets'] = allowed
        session['current_admin_sheet'] = sheet
        return jsonify(success=True, current_sheet=sheet, sheet=sheet)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()

# ============================================================
# Config / Navigator
# ============================================================
@app.route('/api/config', methods=['GET', 'PUT', 'POST'])
def api_config():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if request.method in ['PUT', 'POST']:
            save_config(conn, admin, sheet, get_payload() or {})
        row = get_config_row(conn, admin, sheet)
        return jsonify(serialize_config(row, conn, admin, sheet))
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/event-config', methods=['GET'])
def api_event_config():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        row = get_config_row(conn, admin, sheet)
        return jsonify(success=True, config=serialize_config(row, conn, admin, sheet))
    except Exception as e:
        return jsonify(success=False, message=str(e), config={}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/event-config', methods=['PUT', 'POST'])
@app.route('/api/admin/config', methods=['PUT', 'POST'])
def api_admin_event_config():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        save_config(conn, admin, sheet, get_payload() or {})
        row = get_config_row(conn, admin, sheet)
        return jsonify(success=True, message='設定已儲存', config=serialize_config(row, conn, admin, sheet))
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


# ============================================================
# Products
# ============================================================
@app.route('/api/products', methods=['GET', 'PUT', 'POST'])
def api_products():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if request.method in ['PUT', 'POST']:
            data = get_payload() or {}
            products = data.get('products') if isinstance(data, dict) else []
            if products is None and isinstance(data, list):
                products = data
            save_config(conn, admin, sheet, {'products': products or []})
        cfg = get_config_row(conn, admin, sheet)
        return jsonify(success=True, products=load_products_from_config(cfg))
    except Exception as e:
        return jsonify(success=False, message=str(e), products=[]), 500
    finally:
        if conn:
            conn.close()

# ============================================================
# Agenda / schedule
# ============================================================
@app.route('/api/agenda', methods=['GET', 'PUT', 'POST'])
@app.route('/api/schedule', methods=['GET', 'PUT', 'POST'])
def api_agenda():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if request.method in ['PUT', 'POST']:
            data = get_payload() or {}
            rows = data.get('agenda') or data.get('schedule') or data.get('items') or (data if isinstance(data, list) else [])
            save_agenda(conn, admin, sheet, rows)
        rows = load_agenda(conn, admin, sheet)
        return jsonify(success=True, agenda=rows, schedule=rows, items=rows)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e), agenda=[], schedule=[]), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/admin/schedule', methods=['PUT', 'POST'])
def api_admin_schedule():
    return api_agenda()


# ============================================================
# Industry mappings / exhibitors / companies
# ============================================================
@app.route('/api/industry-mappings', methods=['GET', 'PUT', 'POST'])
def api_industry_mappings():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if request.method in ['PUT', 'POST']:
            data = get_payload() or {}
            rows = data.get('mappings') or data.get('industry_mappings') or data.get('items') or (data if isinstance(data, list) else [])
            save_industry_mappings(conn, admin, sheet, rows)
        rows = load_industry_mappings(conn, admin, sheet)
        return jsonify(success=True, mappings=rows, industry_mappings=rows)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e), mappings=[]), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/exhibitors', methods=['GET', 'PUT', 'POST'])
def api_exhibitors():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if request.method in ['PUT', 'POST']:
            data = get_payload() or {}
            rows = data.get('exhibitors') or data.get('items') or (data if isinstance(data, list) else [])
            save_exhibitors(conn, admin, sheet, rows)
        rows = load_exhibitors(conn, admin, sheet)
        return jsonify(success=True, exhibitors=rows, data=rows)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/companies')
def api_companies():
    admin, sheet = q_event()
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT COALESCE(NULLIF(company,''), company_name) AS company
                FROM event_registrations
                WHERE admin_username=%s AND google_sheet_name=%s
                  AND COALESCE(NULLIF(company,''), company_name, '') <> ''
                ORDER BY company
                """,
                (admin, sheet),
            )
            rows = [r['company'] for r in cur.fetchall() if r.get('company')]
        return jsonify(success=True, companies=rows)
    except Exception as e:
        return jsonify(success=False, message=str(e), companies=[]), 500
    finally:
        if conn:
            conn.close()

# ============================================================
# Registrations / search / check-in
# ============================================================
@app.route('/api/search/<method>')
def api_search(method):
    admin, sheet = q_event()
    method = clean_text(method).lower()
    value = clean_text(
        request.args.get(method)
        or request.args.get('q')
        or request.args.get('value')
        or request.args.get('keyword')
        or request.args.get('query')
        or request.args.get('company')
        or request.args.get('name')
        or request.args.get('phone')
    )
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        if not value:
            return jsonify(success=True, data=[], results=[])

        if method in ['company', 'company_name', 'unit']:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT * FROM event_registrations
                    WHERE admin_username=%s AND google_sheet_name=%s
                    ORDER BY id ASC
                    LIMIT 5000
                    """,
                    (admin, sheet),
                )
                candidates = cur.fetchall()
            scored = []
            for row in candidates:
                score = search_match_score(row, value, method)
                if score > 0:
                    scored.append((score, row))
            scored.sort(key=lambda x: (-x[0], clean_text(x[1].get('company') or x[1].get('company_name')), clean_text(x[1].get('name'))))
            rows = [public_user(r) for _, r in scored[:120]]
            return jsonify(success=True, data=rows, results=rows, count=len(rows))

        like = f"%{value}%"
        if method in ['phone', 'mobile', 'tel']:
            cond = "phone LIKE %s"
            args = [like]
        else:
            cond = "name LIKE %s"
            args = [like]
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT * FROM event_registrations
                WHERE admin_username=%s AND google_sheet_name=%s AND {cond}
                ORDER BY id ASC
                LIMIT 120
                """,
                [admin, sheet] + args,
            )
            rows = [public_user(r) for r in cur.fetchall()]
        return jsonify(success=True, data=rows, results=rows, count=len(rows))
    except Exception as e:
        return jsonify(success=False, message=str(e), data=[], results=[]), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/registrations/add', methods=['POST'])
def api_registration_add():
    conn = None
    try:
        data = get_payload() or {}
        admin = clean_text(data.get('admin')) or session.get('username') or DEFAULT_ADMIN
        sheet = clean_text(data.get('sheet') or data.get('google_sheet_name')) or session.get('current_admin_sheet') or DEFAULT_SHEET
        portrait_bool = bool(data.get('portrait_consent'))
        portrait_status = clean_text(data.get('portrait_consent_status')) or ('同意' if portrait_bool else '不同意')
        now = utc_now_naive()
        conn = get_db_connection()
        ensure_core_tables(conn)
        ensure_config(conn, admin, sheet)
        with conn.cursor() as cur:
            lock_event_mutations(cur, admin, sheet)
            # 具名排除 id，全面精準寫入
            cur.execute(
                """
                INSERT INTO event_registrations
                (admin_username, google_sheet_name, admin_user, event_key, name, phone, email, company, company_name, job_title, industry_category, meal_preference, seat, seating_chart, status, is_original, checked_in_at, checkin_time, portrait_consent, portrait_consent_status, portrait_consent_time, special_notes, note, raw_data)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    admin, sheet, admin, sheet,
                    clean_text(data.get('name')),
                    clean_text(data.get('phone')),
                    clean_text(data.get('email')),
                    clean_text(data.get('company')),
                    clean_text(data.get('company')),
                    clean_text(data.get('job_title')),
                    normalize_industry_category(data.get('industry_category') or data.get('industry')),
                    normalize_meal_preference(data.get('meal_preference') or data.get('meal')),
                    normalize_table_label(data.get('seat')) or '待分桌',
                    normalize_table_label(data.get('seat')) or '待分桌',
                    'checked_in', 1, now, now,
                    1 if portrait_bool else 0,
                    portrait_status,
                    now,
                    clean_text(data.get('special_notes') or data.get('note')),
                    clean_text(data.get('special_notes') or data.get('note')),
                    json_dumps({k: v for k, v in data.items()}),
                ),
            )
            rid = cur.lastrowid
            cur.execute("SELECT * FROM event_registrations WHERE id=%s", (rid,))
            row = public_user(cur.fetchone())
        conn.commit()
        return jsonify(success=True, id=rid, data=row)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/checkin/<int:rid>', methods=['POST'])
def api_checkin(rid):
    conn = None
    try:
        data = get_payload() or {}
        admin, sheet = event_args()
        is_original = data.get('is_original', True)
        proxy = data.get('proxy_info') or data.get('proxy') or {}
        portrait_bool = bool(data.get('portrait_consent'))
        portrait_status = clean_text(data.get('portrait_consent_status') or data.get('image_rights_status')) or ('同意' if portrait_bool else '不同意')
        status = 'checked_in' if is_original else '替代'
        now = utc_now_naive()
        conn = get_db_connection()
        ensure_core_tables(conn)
        ensure_config(conn, admin, sheet)
        with conn.cursor() as cur:
            lock_event_mutations(cur, admin, sheet)
            cur.execute(
                "SELECT * FROM event_registrations "
                "WHERE id=%s AND admin_username=%s AND google_sheet_name=%s LIMIT 1 FOR UPDATE",
                (rid, admin, sheet),
            )
            old = cur.fetchone()
            if not old:
                return jsonify(success=False, message='找不到報到資料'), 404
            updated_seat = normalize_table_label(data.get('seat')) or normalize_table_label(
                old.get('seat') or old.get('seating_chart')
            )
            cur.execute(
                """
                UPDATE event_registrations
                SET status=%s,
                    is_original=%s,
                    proxy_name=%s,
                    proxy_phone=%s,
                    checked_in_at=%s,
                    checkin_time=%s,
                    portrait_consent=%s,
                    portrait_consent_status=%s,
                    portrait_consent_time=%s,
                    seat=%s,
                    seating_chart=%s
                WHERE id=%s AND admin_username=%s AND google_sheet_name=%s
                """,
                (
                    status,
                    1 if is_original else 0,
                    clean_text(proxy.get('name')) if not is_original else '',
                    clean_text(proxy.get('phone')) if not is_original else '',
                    now,
                    now,
                    1 if portrait_bool else 0,
                    portrait_status,
                    now,
                    updated_seat,
                    updated_seat,
                    rid, admin, sheet,
                ),
            )
            if cur.rowcount != 1:
                conn.rollback()
                return jsonify(success=False, message='報到資料已變更，請重新搜尋'), 409
            cur.execute(
                "SELECT * FROM event_registrations "
                "WHERE id=%s AND admin_username=%s AND google_sheet_name=%s",
                (rid, admin, sheet),
            )
            updated = public_user(cur.fetchone())
        conn.commit()
        return jsonify(success=True, data=updated)
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/company-search')
def api_company_search_alias():
    return api_search('company')


@app.route('/api/search')
def api_search_query_alias():
    method = clean_text(request.args.get('method') or request.args.get('type') or 'name')
    return api_search(method)


# ============================================================
# Roster import / export / delete
# ============================================================
@app.route('/api/sheets/import_csv/template', methods=['GET'])
def import_csv_template_api():
    output = io.StringIO(newline='')
    csv.writer(output, lineterminator='\r\n').writerow(REGISTRATION_TEMPLATE_HEADERS)
    data = ('\ufeff' + output.getvalue()).encode('utf-8')
    return Response(
        data,
        content_type='text/csv; charset=utf-8',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote('報到名單匯入範本.csv')}"},
    )


@app.route('/api/sheets/import_csv', methods=['POST'])
def import_csv_api():
    requested_admin, sheet = q_event()
    conn = None
    try:
        mode = clean_text(request.args.get('mode') or request.form.get('mode') or 'preview').lower()
        if mode not in ['preview', 'commit']:
            return jsonify(success=False, message='不支援的名單匯入模式'), 400

        session_username = clean_text(session.get('username'))
        if not session.get('admin_logged_in') or not session_username:
            return jsonify(success=False, message='請先登入管理員後台再匯入名單'), 401
        if requested_admin != session_username:
            return jsonify(success=False, message='你沒有權限操作這個管理員的場次'), 403
        admin = session_username

        upload = request.files.get('file') or request.files.get('csv') or request.files.get('upload')
        if not upload:
            return jsonify(success=False, message='找不到上傳檔案'), 400

        raw = upload.stream.read(MAX_CSV_BYTES + 1)
        if len(raw) > MAX_CSV_BYTES:
            return jsonify(success=False, message='名單檔案超過 5 MB 上限'), 413

        try:
            parsed = parse_registration_upload(raw, upload.filename or '')
        except ValueError as err:
            return jsonify(success=False, message=str(err)), 400

        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT allowed_events FROM admins WHERE username=%s LIMIT 1", (session_username,))
            admin_row = cur.fetchone()
        if not admin_row:
            return jsonify(success=False, message='管理員帳號不存在，請重新登入'), 403
        allowed_sheets = [
            value.strip()
            for value in (admin_row.get('allowed_events') or '').split(',')
            if value.strip()
        ] or [DEFAULT_SHEET]
        if sheet not in allowed_sheets:
            return jsonify(success=False, message='你沒有權限操作這個場次'), 403
        ensure_config(conn, admin, sheet)

        if mode == 'preview':
            preview_token = create_csv_preview_token()
            with conn.cursor() as cur:
                existing_count, roster_revision = load_csv_roster_snapshot(cur, admin, sheet)
                cur.execute(
                    "UPDATE csv_import_previews SET used_at=NOW() "
                    "WHERE admin_username=%s AND google_sheet_name=%s AND used_at IS NULL",
                    (admin, sheet),
                )
                cur.execute(
                    "DELETE FROM csv_import_previews "
                    "WHERE expires_at < DATE_SUB(NOW(), INTERVAL 1 DAY) "
                    "OR used_at < DATE_SUB(NOW(), INTERVAL 1 DAY)"
                )
                cur.execute(
                    """
                    INSERT INTO csv_import_previews
                    (token_hash, admin_username, google_sheet_name, file_sha256,
                     normalized_rows_sha256, source_format, worksheet_name, parser_version,
                     valid_count, roster_count, roster_revision, expires_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,DATE_ADD(NOW(), INTERVAL %s SECOND))
                    """,
                    (
                        csv_preview_token_hash(preview_token), admin, sheet, csv_file_hash(raw),
                        registration_rows_hash(parsed['rows']), parsed['source_format'],
                        parsed.get('worksheet_name'), REGISTRATION_PARSER_VERSION,
                        parsed['valid_count'], existing_count, roster_revision,
                        CSV_PREVIEW_TOKEN_TTL_SECONDS,
                    ),
                )
            conn.commit()

            preview_rows = [
                {
                    'name': row['name'],
                    'phone': row['phone'],
                    'email': row['email'],
                    'company': row['company'],
                    'job_title': row['job_title'],
                    'industry_category': row['industry_category'],
                    'meal_preference': row['meal_preference'],
                    'seat': row['seat'],
                    'special_notes': row['special_notes'],
                }
                for row in parsed['rows'][:10]
            ]
            return jsonify(
                success=True,
                mode='preview',
                message=f"名單驗證完成：可匯入 {parsed['valid_count']} 筆",
                target={'admin': admin, 'sheet': sheet},
                filename=upload.filename or '',
                source_format=parsed['source_format'],
                worksheet_name=parsed.get('worksheet_name'),
                total_count=parsed['total_count'],
                valid_count=parsed['valid_count'],
                skipped_count=parsed['skipped_count'],
                existing_count=existing_count,
                detected_encoding=parsed.get('encoding'),
                headers=parsed['headers'],
                recognized_fields=parsed['recognized_fields'],
                rows=preview_rows,
                preview_token=preview_token,
            )

        preview_token = clean_text(request.form.get('preview_token'))
        if len(preview_token) < 32 or len(preview_token) > 200:
            return jsonify(success=False, message='名單預覽憑證無效，請重新預覽'), 409

        rows = parsed['rows']
        inserted_count = 0
        with conn.cursor() as cur:
            lock_event_mutations(cur, admin, sheet)
            cur.execute(
                """
                SELECT admin_username, google_sheet_name, file_sha256, normalized_rows_sha256,
                       source_format, worksheet_name, parser_version, valid_count,
                       roster_count, roster_revision
                FROM csv_import_previews
                WHERE token_hash=%s AND used_at IS NULL AND expires_at >= NOW()
                FOR UPDATE
                """,
                (csv_preview_token_hash(preview_token),),
            )
            preview = cur.fetchone()
            if not preview:
                conn.rollback()
                return jsonify(success=False, message='名單預覽已失效或已使用，請重新預覽'), 409

            preview_matches = (
                preview.get('admin_username') == admin
                and preview.get('google_sheet_name') == sheet
                and int(preview.get('valid_count') or -1) == parsed['valid_count']
                and hmac.compare_digest(str(preview.get('file_sha256') or ''), csv_file_hash(raw))
                and hmac.compare_digest(
                    str(preview.get('normalized_rows_sha256') or ''),
                    registration_rows_hash(parsed['rows']),
                )
                and preview.get('source_format') == parsed['source_format']
                and clean_text(preview.get('worksheet_name')) == clean_text(parsed.get('worksheet_name'))
                and preview.get('parser_version') == REGISTRATION_PARSER_VERSION
            )
            if not preview_matches:
                conn.rollback()
                return jsonify(success=False, message='名單檔案或目標場次已變更，請重新預覽'), 409

            current_count, current_revision = load_csv_roster_snapshot(cur, admin, sheet, for_update=True)
            if (
                int(preview.get('roster_count') or 0) != current_count
                or not hmac.compare_digest(str(preview.get('roster_revision') or ''), current_revision)
            ):
                conn.rollback()
                return jsonify(success=False, message='名單在預覽後已被更新，請重新預覽再匯入'), 409

            cur.execute(
                "UPDATE csv_import_previews SET used_at=NOW() WHERE token_hash=%s AND used_at IS NULL",
                (csv_preview_token_hash(preview_token),),
            )
            if cur.rowcount != 1:
                conn.rollback()
                return jsonify(success=False, message='名單預覽已失效或已使用，請重新預覽'), 409

            cur.execute("DELETE FROM event_registrations WHERE admin_username=%s AND google_sheet_name=%s", (admin, sheet))
            for r in rows:
                cur.execute(
                    """
                    INSERT INTO event_registrations
                    (admin_username, google_sheet_name, admin_user, event_key, name, phone, email, company, company_name, job_title, region, training_level, industry_category, meal_preference, seat, seating_chart, status, special_notes, note, raw_data)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pending',%s,%s,%s)
                    """,
                    (
                        admin, sheet, admin, sheet,
                        r['name'], r['phone'], r['email'], r['company'], r['company'], r['job_title'],
                        r['region'], r['training_level'], r['industry_category'], r['meal_preference'],
                        r['seat'], r['seat'], r['special_notes'], r['special_notes'], r['raw_data'],
                    ),
                )
                inserted_count += 1
        conn.commit()
        return jsonify(
            success=True,
            mode='commit',
            message=f'名單匯入完成：已寫入 {inserted_count} 筆，略過 {parsed["skipped_count"]} 筆',
            source_format=parsed['source_format'],
            worksheet_name=parsed.get('worksheet_name'),
            count=inserted_count,
            inserted_count=inserted_count,
            total_count=parsed['total_count'],
            valid_count=parsed['valid_count'],
            skipped_count=parsed['skipped_count'],
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/sheets/delete_data', methods=['DELETE', 'POST'])
def delete_sheet_data_api():
    requested_admin, sheet = q_event()
    conn = None
    try:
        admin = clean_text(session.get('username'))
        if not session.get('admin_logged_in') or not admin:
            return jsonify(success=False, message='請先登入管理員後台'), 401
        if requested_admin != admin:
            return jsonify(success=False, message='你沒有權限操作這個管理員的場次'), 403
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT allowed_events FROM admins WHERE username=%s LIMIT 1", (admin,))
            admin_row = cur.fetchone()
        allowed_sheets = [
            value.strip()
            for value in ((admin_row or {}).get('allowed_events') or '').split(',')
            if value.strip()
        ] or [DEFAULT_SHEET]
        if sheet not in allowed_sheets:
            return jsonify(success=False, message='你沒有權限操作這個場次'), 403
        ensure_config(conn, admin, sheet)
        with conn.cursor() as cur:
            lock_event_mutations(cur, admin, sheet)
            cur.execute("DELETE FROM event_registrations WHERE admin_username=%s AND google_sheet_name=%s", (admin, sheet))
        conn.commit()
        return jsonify(success=True, message=f'已成功清空場次「{sheet}」的所有旅客名單資料')
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/sheets/export_csv', methods=['GET'])
def export_csv_api():
    requested_admin, sheet = q_event()
    admin, access_error = validate_admin_event_session(requested_admin, sheet)
    if access_error:
        return access_error
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT * FROM event_registrations
                WHERE admin_username=%s AND google_sheet_name=%s
                ORDER BY id ASC
                """,
                (admin, sheet),
            )
            rows = cur.fetchall()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['原名單姓名', '實際出席姓名', '手機', '公司/單位', 'Email', '職稱', '產業類別', '葷素', '桌號/座位', '報到狀態', '報到時間', '肖像權狀態', '替代人姓名', '替代人手機', '備註'])
        for r in rows:
            user = public_user(r)
            writer.writerow([
                r.get('name', ''), user.get('attendance_name', ''), r.get('phone', ''), r.get('company', ''), r.get('email', ''), r.get('job_title', ''),
                r.get('industry_category', ''), normalize_meal_preference(r.get('meal_preference')), r.get('seat', ''),
                r.get('status', ''), user.get('checkin_time', ''), r.get('portrait_consent_status', ''),
                r.get('proxy_name', ''), r.get('proxy_phone', ''), r.get('special_notes', ''),
            ])
        data = output.getvalue().encode('utf-8-sig')
        filename = re.sub(r'[\\/:*?"<>|\s]+', '_', sheet).strip('_') or 'registrations'
        return Response(
            data,
            content_type='text/csv; charset=utf-8',
            headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}_{taipei_now().strftime('%Y%m%d_%H%M%S')}.csv"},
        )
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500
    finally:
        if conn:
            conn.close()

# ============================================================
# Stats
# ============================================================
def resolve_registration_industry(row, mappings):
    direct = normalize_industry_category((row or {}).get('industry_category'))
    if direct:
        return direct
    company = clean_text((row or {}).get('company') or (row or {}).get('company_name'))
    company_norm = search_norm(company)
    for item in mappings or []:
        keyword = clean_text(item.get('keyword') or item.get('company_name') or item.get('company'))
        if keyword and search_norm(keyword) in company_norm:
            return clean_text(item.get('category') or item.get('industry')) or '其他'
    return '未分類'


def aggregate_registration_field(rows, value_getter):
    counter = {}
    for row in rows or []:
        label = clean_text(value_getter(row)) or '未填'
        counter[label] = counter.get(label, 0) + 1
    total = sum(counter.values())
    return [
        {
            'name': name,
            'count': count,
            'percent': round((count / total) * 100, 1) if total else 0,
        }
        for name, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]


def aggregate_checked_in_companies(rows):
    companies = {}
    for raw_row in rows or []:
        if not status_checked((raw_row or {}).get('status')):
            continue
        row = public_user(raw_row)
        company = clean_text(row.get('company') or row.get('company_name'))
        if not company:
            continue
        key = search_norm(company)
        item = companies.setdefault(key, {
            'name': company,
            'company_name': company,
            'checked_in_count': 0,
            'attendees': [],
            'industries': [],
        })
        item['checked_in_count'] += 1
        attendee = clean_text(row.get('display_name') or row.get('name'))
        if attendee and attendee not in item['attendees']:
            item['attendees'].append(attendee)
        industry = normalize_industry_category(row.get('industry_category'))
        if industry and industry not in item['industries']:
            item['industries'].append(industry)
    return sorted(companies.values(), key=lambda item: (-item['checked_in_count'], item['name']))


@app.route('/api/dashboard_stats')
@app.route('/api/admin/dashboard_stats')
def dashboard_stats():
    requested_admin, sheet = q_event()
    include_private = request.path.startswith('/api/admin/')
    if include_private:
        admin, access_error = validate_admin_event_session(requested_admin, sheet)
        if access_error:
            return access_error
    else:
        admin = requested_admin
    conn = None
    try:
        conn = get_db_connection()
        ensure_core_tables(conn)
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN status IN ('checked_in','已報到','替代','done') THEN 1 ELSE 0 END) AS checked
                FROM event_registrations
                WHERE admin_username=%s AND google_sheet_name=%s
                """,
                (admin, sheet),
            )
            s = cur.fetchone() or {}
            total = int(s.get('total') or 0)
            checked = int(s.get('checked') or 0)

            cur.execute(
                """
                SELECT id, name, phone, email, company, company_name, job_title,
                       industry_category, meal_preference,
                       COALESCE(NULLIF(seat,''), seating_chart, '未分桌') AS seat,
                       status, is_original, proxy_name, proxy_phone,
                       checked_in_at, checkin_time, portrait_consent_status, special_notes
                FROM event_registrations
                WHERE admin_username=%s AND google_sheet_name=%s
                ORDER BY id ASC
                """,
                (admin, sheet),
            )
            roster_rows = cur.fetchall()
            table_detail_map = {}
            for raw_row in roster_rows:
                row = public_user(raw_row)
                seat_val = clean_text(row.get('seat') or '未分桌')
                if not seat_val:
                    seat_val = '未分桌'
                
                table_val = seat_val.split('-')[0].strip() if '-' in seat_val else seat_val
                table_detail_map.setdefault(table_val, []).append(row)

            table_stats = []
            table_details = []
            for table_val, members in sorted(table_detail_map.items(), key=lambda kv: table_sort_key(kv[0])):
                checked_count = sum(1 for m in members if status_checked(m.get('status')))
                vegetarian_total = sum(1 for m in members if is_vegetarian_meal(m.get('meal_preference')))
                vegetarian_checked_in = sum(
                    1 for m in members
                    if status_checked(m.get('status')) and is_vegetarian_meal(m.get('meal_preference'))
                )
                total_count = len(members)
                table_stats.append({
                    'table': table_val,
                    'seat': table_val,
                    'total': total_count,
                    'checked_in': checked_count,
                    'percent': round((checked_count / total_count) * 100, 1) if total_count else 0,
                    'vegetarian_total': vegetarian_total,
                    'vegetarian_checked_in': vegetarian_checked_in,
                })
                table_details.append({
                    'table': table_val,
                    'seat': table_val,
                    'members': members,
                    'checked_in': checked_count,
                    'total': total_count,
                    'vegetarian_total': vegetarian_total,
                    'vegetarian_checked_in': vegetarian_checked_in,
                })

        mappings = load_industry_mappings(conn, admin, sheet)
        checked_rows = [row for row in roster_rows if status_checked(row.get('status'))]
        checked_in_companies = aggregate_checked_in_companies(roster_rows)
        if not include_private:
            checked_in_companies = [
                {key: value for key, value in company.items() if key != 'attendees'}
                for company in checked_in_companies
            ]
        industry_stats = {
            'all': aggregate_registration_field(
                roster_rows, lambda row: resolve_registration_industry(row, mappings)
            ),
            'checked_in': aggregate_registration_field(
                checked_rows, lambda row: resolve_registration_industry(row, mappings)
            ),
        }
        meal_stats = {
            'all': aggregate_registration_field(
                roster_rows, lambda row: normalize_meal_preference(row.get('meal_preference'))
            ),
            'checked_in': aggregate_registration_field(
                checked_rows, lambda row: normalize_meal_preference(row.get('meal_preference'))
            ),
        }
        logs = get_logs(conn, admin, sheet, 25, checked_only=True) if include_private else []
        return jsonify(success=True, stats={
            'total': total,
            'checked_in': checked,
            'not_checked_in': max(total - checked, 0),
            'logs': logs,
            'industry_stats': industry_stats,
            'meal_stats': meal_stats,
            'checked_in_companies': checked_in_companies,
            'table_stats': table_stats,
            'table_details': table_details if include_private else [],
        })
    except Exception as e:
        return jsonify(success=False, message=str(e), stats={}), 500
    finally:
        if conn:
            conn.close()

if __name__ == '__main__':
    port = int(os.getenv('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=os.getenv('FLASK_DEBUG') == '1')
